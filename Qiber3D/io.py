import json
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from textwrap import dedent
from zipfile import ZipFile

import networkx as nx
import numpy as np
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from scipy import interpolate

import Qiber3D
from Qiber3D import helper, config


class IO:

    class load:
        def __new__(cls, path, **kwargs):
            """
            Returns a new :class:`Qiber3D.Network` from file.

            Supports: :file:`.qiber`, :file:`.json`, :file:`.mv3d`, :file:`.tif`, :file:`.nd2`, :file:`.swc`, :file:`.ntr`

            :param path: file path to load
            :type path: str, Path
            :param kwargs: key-word arguments are passed down to the individual IO functions
            """

            path = Path(path)
            if path.suffix == '':
                path = path.with_suffix('.qiber')

            if path.suffix == '.qiber':
                return cls.binary(path)
            elif path.suffix == '.json':
                return cls.json(path)
            elif path.suffix == '.mv3d':
                return cls.mv3d(path)
            elif path.suffix == '.nd2':
                return cls.nd2(path, **kwargs)
            elif path.suffix == '.ntr':
                return cls.ntr(path)
            elif path.suffix == '.swc':
                return cls.swc(path, **kwargs)
            else:
                return cls.image(path, **kwargs)

        @staticmethod
        def binary(path):
            """
            Create a :class:`Qiber3D.Network` from a :file:`.qiber` file, created by :meth:`Qiber3D.Network.save`

            :param path: file path to load
            :type path: str, Path
            :return: :class:`Qiber3D.Network`
            """
            path = Path(path)
            with ZipFile(path, mode='r') as save_file:
                net = IO.load.json(path, data=json.loads(save_file.read('network.json').decode('utf-8')))
                try:
                    extractor_steps = helper.NumpyMemoryManager.load(
                        fileobj=BytesIO(save_file.read('extractor_steps.tar')))
                except KeyError:
                    extractor_steps = None
            net.extractor_steps = extractor_steps
            return net

        @staticmethod
        def image(path, channel=None, voxel_size=None):
            """
            Create a :class:`Qiber3D.Network` from a image file.

            :param path: file path to load
            :type path: str, Path
            :param channel: either index or name of image channel
            :type channel: int, str
            :param tuple(float) voxel_size: physical size of voxel in (x,y,z)
            :return: :class:`Qiber3D.Network`

            """
            path = Path(path)

            if voxel_size is None:
                voxel_size = config.extract.voxel_size

            if voxel_size is None:
                print('Please set the physical size of the voxels')
                vsx = float(input('\tx: '))
                vsy = float(input('\ty: '))
                vsz = float(input('\tz: '))
                voxel_size = (vsx, vsy, vsz)
            ex = Qiber3D.Extractor(path, channel=channel, voxel_size=voxel_size)
            return ex.get_network()

        @staticmethod
        def nd2(path, channel=None):
            """
            Create a :class:`Qiber3D.Network` from a :file:`.nd2` file.

            :param path: file path to load
            :type path: str, Path
            :param channel: either index or name of image channel
            :type channel: int, str
            :return: :class:`Qiber3D.Network`
            """
            path = Path(path)
            ex = Qiber3D.Extractor(path, channel=channel)
            return ex.get_network()

        @staticmethod
        def mv3d(path):
            """
            Create a :class:`Qiber3D.Network` from a :file:`.mv3d` file.

            :param path: file path to load
            :type path: str, Path
            :return: :class:`Qiber3D.Network`
            """
            path = Path(path)

            network_data = np.loadtxt(path, delimiter='\t')
            segments = {}
            available_segments = [int(sid) for sid in (set(network_data[:, 0].astype(int)))]

            for seg_id in available_segments:
                segments[seg_id] = dict(
                    seg_id=seg_id,
                    points=network_data[network_data[:, 0] == seg_id][:, 1:4],
                    radius=network_data[network_data[:, 0] == seg_id][:, 4] / 2.0
                )
            data = {
                'path': path,
                'name': path.with_suffix('').name,
                'segments': segments}
            return Qiber3D.Network(data)

        @staticmethod
        def network(net, scale=1, input_path=None, segment_list=None):
            """
            Create a new :class:`Qiber3D.Network` from a :class:`Qiber3D.Network`.

            :param Qiber3D.Network net: original network
            :param float scale: scale all points and radii by this value
            :param input_path: set this path as new imput path of the returning network
            :type input_path: str, Path
            :param tuple segment_list: limit the new network to this list of segments (sid)
            :return: :class:`Qiber3D.Network`
            """

            segments = {}
            if segment_list is None:
                for seg_id, segment in enumerate(net.segment.values()):
                    segments[seg_id] = {
                        'seg_id': seg_id,
                        'points': segment.point * scale,
                        'radius': segment.radius * scale
                    }
            else:
                for new_seg_id, old_seg_id in enumerate(segment_list):
                    segments[new_seg_id] = {
                        'seg_id': new_seg_id,
                        'points': net.segment[old_seg_id].point * scale,
                        'radius': net.segment[old_seg_id].radius * scale
                    }
            if input_path is not None:
                data = {
                    'path': input_path,
                    'name': input_path.with_suffix('').name,
                    'segments': segments}
            else:
                data = {
                    'path': net.input_file,
                    'name': net.name,
                    'segments': segments}
            return Qiber3D.Network(data)

        @staticmethod
        def ntr(path):
            """
            Create a :class:`Qiber3D.Network` from a :file:`.ntr` file.

            :param path: file path to load
            :type path: str, Path
            :return: :class:`Qiber3D.Network`
            """
            path = Path(path)
            content = path.read_text()
            data_re = re.compile(r'(.*),(.*),(.*),(.*),(.*),(.*)\n')

            search = data_re.findall(content)
            data = np.array(search, dtype=np.float32)

            seg_id = 0
            segments = {seg_id: {'seg_id': seg_id, 'points': [], 'radius': []}}
            unfinished_bp = []
            for n, (x, y, z, d, pt, pd) in enumerate(data):
                segments[seg_id]['points'].append((x, y, z))
                segments[seg_id]['radius'].append(d/2.0)
                if int(pt) in [7, 8]:
                    seg_id += 1
                    segments[seg_id] = {'seg_id': seg_id, 'points': [], 'radius': []}
                if int(pt) == 7:
                    unfinished_bp.append((x, y, z, d))
                    segments[seg_id]['points'].append((x, y, z))
                    segments[seg_id]['radius'].append(d / 2.0)
                if int(pt) == 8:
                    if unfinished_bp:
                        x, y, z, d = unfinished_bp.pop()
                        segments[seg_id]['points'].append((x, y, z))
                        segments[seg_id]['radius'].append(d / 2.0)
                    pass

            to_remove = []
            for seg_id in segments:
                if segments[seg_id]['points']:
                    segments[seg_id]['points'] = np.array(segments[seg_id]['points'])
                    segments[seg_id]['radius'] = np.array(segments[seg_id]['radius'])
                else:
                    to_remove.append(seg_id)
                if len(segments[seg_id]['points']) == 2:
                    if np.all(segments[seg_id]['points'][0] == segments[seg_id]['points'][1]):
                        to_remove.append(seg_id)
            for seg_id in to_remove:
                segments.__delitem__(seg_id)

            data = {
                'path': path,
                'name': path.with_suffix('').name,
                'segments': segments}
            return Qiber3D.Network(data)

        @staticmethod
        def _from_graph(graph, point_lookup, radius_lookup, scale=1.0, ravel=False):
            segment_data = {}
            segments = []
            for fiber in (graph.subgraph(c).copy() for c in nx.connected_components(graph)):
                start_points = None
                stop_points = []
                for node in fiber:
                    if len(fiber.adj[node]) == 1:
                        start_points = [(node,)]
                        break
                if start_points is None:
                    start_node = list(fiber.nodes)[0]
                    start_points = [list(graph.edges(start_node))[0]]

                while start_points:
                    start = start_points.pop()
                    if len(start) == 1:
                        new_segment = [start[0]]
                        start = start[0]
                    else:
                        if not fiber.has_edge(start[0], start[1]):
                            continue
                        new_segment = [start[0], start[1]]
                        fiber.remove_edge(start[0], start[1])
                        start = start[1]
                    f = start
                    while True:
                        t = list(fiber.neighbors(f))
                        if len(t) == 1:
                            new_segment.append(t[0])
                            fiber.remove_edge(f, t[0])
                            f = t[0]
                            if t[0] in stop_points:
                                break
                        elif len(t) > 1:
                            for paths in t:
                                start_points.append((f, paths))
                            stop_points.append(f)
                            break
                        else:
                            break
                    segments.append(new_segment)

            for sid, seg in enumerate(segments):
                if ravel:
                    points = np.array(np.unravel_index(seg, point_lookup)).T * scale
                    points[:] = points[:, (2, 1, 0)]
                    radius = radius_lookup.flat[seg] * scale
                else:
                    points = [point_lookup[pid] for pid in seg]
                    radius = np.array([radius_lookup[pid]*scale for pid in seg])
                    points = np.array([(round(x*scale, 4), round(y*scale, 4), round(z*scale, 4)) for (x, y, z) in points])

                segment_data[sid] = dict(
                    points=points,
                    radius=radius,
                    seg_id=sid
                )
            return segment_data

        @staticmethod
        def swc(path, allowed_types=None):
            """
            Create a :class:`Qiber3D.Network` from a :file:`.swc` file.

            :param path: file path to load
            :type path: str, Path
            :param tuple allowed_types: limit the returned network to these SEGMENT_TYPEs
            :return: :class:`Qiber3D.Network`
            """

            path = Path(path)
            raw_data = np.loadtxt(path, comments='#')
            point_lookup = helper.PointLookUp(places=1)
            id_mapper = {}
            radius_lookup = {}
            graph = nx.Graph()
            for n, pi, x, y, z, r, parent in raw_data:
                if allowed_types is not None:
                    if int(pi) not in allowed_types:
                        continue
                pid = point_lookup[(x, y, z)]
                id_mapper[int(n)] = pid
                radius_lookup[pid] = r
                if int(parent) != -1 and int(parent) in id_mapper:
                    graph.add_edge(pid, id_mapper[int(parent)])
            segment_data = IO.load._from_graph(graph, point_lookup, radius_lookup)
            data = {
                'path': path,
                'name': path.with_suffix('').name,
                'segments': segment_data}
            return Qiber3D.Network(data)

        @staticmethod
        def json(path, data=None):
            """
            Create a :class:`Qiber3D.Network` from a :file:`.json` file.

            :param path: file path to load
            :type path: str, Path
            :param dict data: load network from a ``dict`` representation of the :file:`.json` file directly
            :return: :class:`Qiber3D.Network`
            """
            path = Path(path)
            if data is None:
                data = json.loads(path.read_text())
            segment_data = {}
            for segment in data['network']['segment'].values():
                segment_data[segment['sid']] = dict(
                    seg_id=segment['sid'],
                    points=np.array(segment['point']),
                    radius=np.array(segment['radius'])
                )
            network_data = {
                'path': path,
                'name': data['meta']['name'],
                'segments': segment_data}
            net = Qiber3D.Network(network_data)
            net.extractor_data = data['network'].get('extractor_data')
            return net

        @staticmethod
        def __fill_in_segment(segment, resolution=200):
            base_points = np.array(segment).T
            tck, u = interpolate.splprep(base_points)
            u_fine = np.linspace(0, 1, resolution)
            x, y, z, r = interpolate.splev(u_fine, tck)
            points = np.stack((x, y, z)).T
            return points, r

        @classmethod
        def synthetic_network(cls):
            """
            Create the synthetic test network.

            :return: :class:`Qiber3D.Network`
            """
            resolution = 200
            raw_segments = []

            # Tree
            tree_start = ((200, 100, 0, 5), (300, 500, 100, 8), (600, 600, 250, 15),
                          (800, 800, 300, 10), (1100, 900, 480, 6.5), (1300, 1000, 500, 6))
            tree_start = cls.__fill_in_segment(tree_start, resolution)
            raw_segments.append(tree_start)
            x, y, z = tree_start[0][2 * resolution // 10]
            r = tree_start[1][2 * resolution // 10]
            tree_first_branch = ((x, y, z, r), (400, 300, 150, 5), (600, 200, 180, 8),
                                 (900, 100, 200, 5), (1280, 110, 400, 4))
            raw_segments.append(cls.__fill_in_segment(tree_first_branch, resolution))
            x, y, z = tree_start[0][4 * resolution // 10]
            r = tree_start[1][4 * resolution // 10]
            tree_second_branch = ((x, y, z, r), (800, 500, 260, 18), (1200, 300, 270, 17),
                                  (1500, 600, 290, 10), (1500, 900, 210, 9))
            raw_segments.append(cls.__fill_in_segment(tree_second_branch, resolution))

            # Tree with loop
            tree_loop_start = ((100, 1100, 300, 5), (200, 950, 330, 13), (900, 800, 250, 20),
                               (950, 600, 200, 22), (1100, 200, 100, 8), (1310, 90, 50, 6))
            tree_loop_start = cls.__fill_in_segment(tree_loop_start, resolution)
            raw_segments.append(tree_loop_start)
            s_x, s_y, s_z = tree_loop_start[0][1 * resolution // 10]
            s_r = tree_loop_start[1][1 * resolution // 10]
            f_x, f_y, f_z = tree_loop_start[0][9 * resolution // 10]
            f_r = tree_loop_start[1][9 * resolution // 10]
            tree_loop_branch = ((s_x, s_y, s_z, s_r), (600, 500, 250, 6), (800, 400, 130, 5), (f_x, f_y, f_z, f_r))
            raw_segments.append(cls.__fill_in_segment(tree_loop_branch, resolution))

            # No branch point
            no_branch = ((1500, 100, 450, 6), (1500, 200, 500, 11), (1400, 300, 500, 14), (1450, 500, 450, 11),
                         (1300, 800, 300, 14), (1200, 900, 0, 9), (1100, 1100, 50, 8))
            raw_segments.append(cls.__fill_in_segment(no_branch, resolution))

            # Circle
            circle = ((780, 700, 428.8, 8), (650, 200, 250, 12), (600, 400, 100, 15),
                      (700, 800, 100, 10), (750, 1100, 250, 7), (780, 700, 428.8, 8))
            raw_segments.append(cls.__fill_in_segment(circle, resolution))

            node_lookup = helper.PointLookUp(places=0)
            radius_lookup = {}
            raw_network = nx.Graph()
            for sid, (points, radii) in enumerate(raw_segments):
                last_node = None
                for point, r in zip(points, radii):
                    pid = node_lookup[point]
                    radius_lookup[pid] = r
                    if last_node is not None:
                        raw_network.add_edge(last_node, pid)
                    last_node = pid

            segment_data = cls._from_graph(raw_network, node_lookup, radius_lookup, scale=0.1)

            data = {
                'path': None,
                'name': 'synthetic',
                'segments': segment_data}
            return Qiber3D.Network(data)


    class export:

        def __new__(cls, net, out_path='.', overwrite=False, mode=None, **kwargs):
            """
            Export a :class:`Qiber3D.Network` to file. Selecting the appropriate format based on the file suffix.

            Supports: :file:`.qiber`, :file:`.json`, :file:`.mv3d`, :file:`.tif`, :file:`.nd2`, :file:`.swc`,
            :file:`.csv`,  :file:`.tsv`, :file:`.xlsx`, :file:`.x3d`

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :param str mode: select the file format ignoring the file suffix.
                Choose from ['binary', 'json', 'mv3d', 'x3d', 'swc', 'xlsx', 'csv', 'tsv', 'tif']
            :param kwargs: key-word arguments are passed down to the individual IO functions
            :return: path to saved file
            :rtype: Path
            """
            if mode is None:
                path = Path(out_path)
                if path.suffix == '.json':
                    mode = 'json'
                elif path.suffix == '.qiber':
                    mode = 'binary'
                elif path.suffix == '.mv3d':
                    mode = 'mv3d'
                elif path.suffix == '.x3d':
                    mode = 'x3d'
                elif path.suffix == '.swc':
                    mode = 'swc'
                elif path.suffix == '.xlsx':
                    mode = 'xlsx'
                elif path.suffix == '.csv':
                    mode = 'csv'
                elif path.suffix == '.tsv':
                    mode = 'tsv'
                elif path.suffix in ['.tif', '.tiff']:
                    mode = 'tif'
                else:
                    mode = 'binary'

            if mode == 'binary':
                return cls.binary(net, out_path=out_path, overwrite=overwrite)
            elif mode == 'json':
                return cls.json(net, out_path=out_path, overwrite=overwrite)
            elif mode == 'mv3d':
                return cls.mv3d(net, out_path=out_path, overwrite=overwrite)
            elif mode == 'swc':
                return cls.swc(net, out_path=out_path, overwrite=overwrite, **kwargs)
            elif mode == 'x3d':
                return cls.x3d(net, out_path=out_path, overwrite=overwrite, **kwargs)
            elif mode == 'xlsx':
                return cls.xlsx(net, out_path=out_path, overwrite=overwrite)
            elif mode == 'csv':
                return cls.csv(net, out_path=out_path, overwrite=overwrite, **kwargs)
            elif mode == 'tsv':
                return cls.csv(net, out_path=out_path, overwrite=overwrite, separator='\t')
            elif mode == 'tif':
                return cls.tif(net, out_path=out_path, overwrite=overwrite, **kwargs)
            else:
                net.logger.warn(f'Could not find mode "{mode}"')
                return None

        @staticmethod
        def binary(net, out_path='.', overwrite=False, save_steps=False):
            """
            Export :class:`Qiber3D.Network` as binary file (:file:`.qiber`).

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :param bool save_steps: save extraction steps image stacks
            :return: path to saved file
            :rtype: Path
            """
            out_path, needs_unlink = helper.out_path_check(out_path, network=net, prefix='', suffix='.qiber',
                                                           overwrite=overwrite, logger=net.logger)
            if out_path is None:
                return

            work_dir = Path(out_path.with_suffix(''))
            work_dir.mkdir(parents=True, exist_ok=True)

            add_paths = [IO.export.json(net, out_path=work_dir / 'network.json', overwrite=True)]
            if save_steps:
                if isinstance(net.extractor_steps, helper.NumpyMemoryManager):
                    net.extractor_steps.save(work_dir / 'extractor_steps.tar')
                    add_paths.append(work_dir / 'extractor_steps.tar')
            with ZipFile(out_path, mode='w') as save_file:
                for add_path in add_paths:
                    save_file.write(add_path, arcname=add_path.name)
                    add_path.unlink()
            work_dir.rmdir()

            return out_path

        @staticmethod
        def xlsx(net, out_path='.', overwrite=False):
            """
            Export :class:`Qiber3D.Network` as Excel file (:file:`.xlsx`).

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :return: path to saved file
            :rtype: Path
            """
            out_path, needs_unlink = helper.out_path_check(out_path, network=net, prefix='', suffix='.xlsx',
                                                           overwrite=overwrite, logger=net.logger)
            if out_path is None:
                return

            net_properties = {
                'average_radius': 'Average radius',
                'max_radius': 'Max radius',
                'cylinder_radius': 'Equal cylinder radius',
                'length': 'Length',
                'volume': 'Volume',
                'bbox_volume': 'Bounding box volume',
                'bbox': 'Bounding box',
                'bbox_size': 'Bounding box size',
                'center': 'Bounding box center'
            }
            fiber_seg_properties = {
                'average_radius': 'Average radius',
                'max_radius': 'Max radius',
                'cylinder_radius': 'Equal cylinder radius',
                'length': 'Length',
                'volume': 'Volume',
                # 'raster_volume',
            }

            wb = Workbook(write_only=True)
            ws = wb.create_sheet('Network')
            title = WriteOnlyCell(ws, f'{net.name}')
            ws.column_dimensions['A'].width = 21
            ws.column_dimensions['B'].width = 21
            title.style = 'Title'
            ws.append([title])
            ws.append([config.app_name, config.version])
            ws.append([])
            subtitle = WriteOnlyCell(ws, 'Metadata')
            empty_subtitle = WriteOnlyCell(ws, '')
            subtitle.style = 'Headline 3'
            empty_subtitle.style = 'Headline 3'
            ws.append([subtitle, empty_subtitle])
            if isinstance(net.input_file, Path):
                ws.append(['Source file', str(net.input_file.absolute())])
            else:
                ws.append(['Source file', '-'])
            ws.append(['Creation date', datetime.now()])

            ws.append([])
            subtitle = WriteOnlyCell(ws, 'Network measurements')
            empty_subtitle = WriteOnlyCell(ws, '')
            subtitle.style = 'Headline 3'
            empty_subtitle.style = 'Headline 3'
            ws.append([subtitle, empty_subtitle])
            ws.append(['Number of fibers', len(net.fiber)])
            ws.append(['Number of segments', len(net.segment)])
            ws.append(['Number of points', len(net.point)])
            ws.append(['Number of branch points', len(net.cross_point_dict)])
            for key, desciption in net_properties.items():
                value = getattr(net, key)
                if type(value) == np.ndarray:
                    value = str(value.tolist())
                if isinstance(value, (np.floating, float)):
                    value = WriteOnlyCell(ws, value=value)
                    value.number_format = '0.00'
                ws.append([desciption, value])

            ws = wb.create_sheet('Fibers')
            ws.column_dimensions['A'].width = 21
            ws.column_dimensions['B'].width = 21
            for fid, fiber in net.fiber.items():
                subtitle = WriteOnlyCell(ws, f'Fiber {fid} measurements')
                empty_subtitle = WriteOnlyCell(ws, '')
                subtitle.style = 'Headline 3'
                empty_subtitle.style = 'Headline 3'
                ws.append([subtitle, empty_subtitle])
                ws.append(['Number of segments', len(fiber.segment)])
                ws.append(['Number of points', sum([len(seg) for seg in fiber.segment.values()])])
                branch_points_raw = sum((list(a) for a in fiber.graph.edges), [])
                check = []
                bp_set = set()
                for bp in branch_points_raw:
                    if bp in check:
                        bp_set.add(bp)
                    else:
                        check.append(bp)
                ws.append(['Number of branch points', len(bp_set)])
                for key, desciption in fiber_seg_properties.items():
                    value = getattr(net, key)
                    if type(value) == np.ndarray:
                        value = str(value.tolist())
                    if isinstance(value, (np.floating, float)):
                        value = WriteOnlyCell(ws, value=value)
                        value.number_format = '0.00'
                    ws.append([desciption, value])
                ws.append(['Segment list'] + [sid for sid in fiber.segment.keys()])
                ws.append([])

            ws = wb.create_sheet('Segments')
            ws.column_dimensions['A'].width = 21
            ws.column_dimensions['B'].width = 21
            for sid, segment in net.segment.items():
                subtitle = WriteOnlyCell(ws, f'Segment {sid} measurements')
                empty_subtitle = WriteOnlyCell(ws, '')
                subtitle.style = 'Headline 3'
                empty_subtitle.style = 'Headline 3'
                ws.append([subtitle, empty_subtitle])
                ws.append(['Number of points', len(segment)])
                for key, desciption in fiber_seg_properties.items():
                    value = getattr(net, key)
                    if type(value) == np.ndarray:
                        value = str(value.tolist())
                    if isinstance(value, (np.floating, float)):
                        value = WriteOnlyCell(ws, value=value)
                        value.number_format = '0.00'
                    ws.append([desciption, value])
                ws.append([])

            ws = wb.create_sheet('Points')
            ws.append(['FID', 'SID', 'X', 'Y', 'Z', 'Radius'])
            for fid, fiber in net.fiber.items():
                for sid, segment in fiber.segment.items():
                    for n, (x, y, z) in enumerate(segment.point):
                        x = WriteOnlyCell(ws, value=x)
                        x.number_format = '0.000'
                        y = WriteOnlyCell(ws, value=y)
                        y.number_format = '0.000'
                        z = WriteOnlyCell(ws, value=z)
                        z.number_format = '0.000'
                        r = WriteOnlyCell(ws, value=segment.radius[n])
                        r.number_format = '0.000'
                        ws.append([fid, sid, x, y, z, r])

            wb.save(out_path)
            return out_path

        @staticmethod
        def csv(net, out_path='.', overwrite=False, separator=';'):
            """
            Export :class:`Qiber3D.Network` as :file:`.csv` file.

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :param str separator: char to separate values
            :return: path to saved file
            :rtype: Path
            """
            out_path, needs_unlink = helper.out_path_check(out_path, network=net, prefix='', suffix='.csv',
                                                           overwrite=overwrite, logger=net.logger)
            if out_path is None:
                return
            out_text = separator.join(('FID', 'SID', 'X', 'Y', 'Z', 'Radius')) + '\n'
            for fid, fiber in net.fiber.items():
                for sid, segment in fiber.segment.items():
                    for n, (x, y, z) in enumerate(segment.point):
                        out_text += separator.join([str(fid), str(sid), f'{x:.3f}', f'{y:.3f}', f'{z:.3f}',
                                                    f'{segment.radius[n]:.3f}']) + '\n'

            out_path.write_text(out_text)
            return out_path

        @staticmethod
        def swc(net, out_path='.', overwrite=False, multiple_files=False):
            """
            Export :class:`Qiber3D.Network` as :file:`.swc` file.

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :param bool multiple_files: save each fiber as separate swc file
            :return: path to saved file
            :rtype: Path
            """
            out_path, needs_unlink = helper.out_path_check(out_path, network=net, prefix='', suffix='.swc',
                                                           overwrite=overwrite, logger=net.logger)
            if out_path is None:
                return
            if isinstance(net.input_file, Path):
                raw = str(net.input_file.absolute())
            else:
                raw = str(net.input_file)
            header = dedent(f"""\
            # ORIGINAL_SOURCE created by {config.app_name} {config.version}
            # CREATURE
            # REGION
            # FIELD/LAYER
            # TYPE {net.name}
            # CONTRIBUTOR
            # REFERENCE      
            # RAW  {raw}          
            # EXTRAS         
            # SOMA_AREA      
            # SHRINKAGE_CORRECTION
            # VERSION_NUMBER      
            # VERSION_DATE 
            # SCALE 1.0 1.0 1.0 micrometers
            # SEGMENT_TYPE    1	      {net.cylinder_radius:.2f}     UNSPECIFIC
            """)

            graph = nx.Graph()
            point_lookup = helper.PointLookUp()
            radius_lookup = {}
            center= net.center
            for segment in net.segment.values():
                last_pid = None
                for n, raw_point in enumerate(segment.point):
                    point = raw_point - center
                    if point in point_lookup:
                        current_pid = point_lookup[point]
                    else:
                        current_pid = point_lookup[point]
                        radius_lookup[current_pid] = segment.radius[n]
                    if last_pid is not None:
                        graph.add_edge(current_pid, last_pid)
                    last_pid = current_pid

            network_data = []
            for fiber in (graph.subgraph(c).copy() for c in nx.connected_components(graph)):
                fiber_data = []
                start_points = None
                stop_points = []
                for node in fiber:
                    if len(fiber.adj[node]) == 1:
                        start_points = [(node,)]
                        break
                if start_points is None:
                    start_points = [(list(fiber.nodes)[0],)]
                while start_points:
                    start = start_points.pop()
                    if len(start) == 1:
                        x, y, z = point_lookup[start[0]]
                        fiber_data.append((start[0], 1, x, y, z, radius_lookup[start[0]], -1))
                        start = start[0]
                    else:
                        if not fiber.has_edge(start[0], start[1]):
                            continue
                        x, y, z = point_lookup[start[1]]
                        fiber_data.append((start[1], 1, x, y, z, radius_lookup[start[1]], start[0]))
                        fiber.remove_edge(start[0], start[1])
                        start = start[1]
                    for f, t in nx.dfs_successors(fiber, start).items():
                        if len(t) == 1:
                            x, y, z = point_lookup[t[0]]
                            fiber_data.append((t[0], 1, x, y, z, radius_lookup[t[0]], f))
                            fiber.remove_edge(f, t[0])
                            if t[0] in stop_points:
                                break
                        elif len(t) > 1:
                            for paths in t:
                                start_points.append((f, paths))
                            stop_points.append(f)
                            break
                network_data.append(fiber_data)

            if multiple_files:
                base_dir = out_path.with_suffix('')
                base_dir.mkdir(exist_ok=True)
                for n, point_list in enumerate(network_data):
                    part_out_path = base_dir / f'{out_path.stem}_{n+1:04}{out_path.suffix}'
                    local_point_lookup = {-1: -1}
                    for pid, entry in enumerate(point_list):
                        local_point_lookup[entry[0]] = pid + 1
                    point_text = [f'{local_point_lookup[entry[0]]:d} {entry[1]:4d} {entry[2]:8.1f} {entry[3]:8.1f} '
                                  f'{entry[4]:8.1f} {entry[5]:6.2f} {local_point_lookup[entry[6]]:4d}'
                                  for entry in point_list]
                    out_text = header + '\n'.join(point_text)
                    part_out_path.write_text(out_text)
            else:
                offset = 0
                point_text = []
                for n, point_list in enumerate(network_data):
                    local_point_lookup = {-1: -1}
                    for pid, entry in enumerate(point_list):
                        local_point_lookup[entry[0]] = pid + 1 + offset
                    point_text += [f'{local_point_lookup[entry[0]]:d} {entry[1]:4d} {entry[2]:8.1f} '
                                   f'{entry[3]:8.1f} {entry[4]:8.1f} {entry[5]:6.2f} {local_point_lookup[entry[6]]:4d}'
                                   for entry in point_list]
                    offset = len(point_list)
                out_text = header + '\n'.join(point_text)
                out_path.write_text(out_text)

            return out_path

        @staticmethod
        def mv3d(net, out_path='.', overwrite=False):
            """
            Export :class:`Qiber3D.Network` as :file:`.mv3d` file.

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :return: path to saved file
            :rtype: Path
            """
            out_path, needs_unlink = helper.out_path_check(out_path, network=net, prefix='', suffix='.mv3d',
                                                           overwrite=overwrite, logger=net.logger)
            if out_path is None:
                return

            number_of_points = 0
            data_part = ""
            for seg in net.segment.values():
                seg_data = []
                for n, (x, y, z) in enumerate(seg.point):
                    seg_data.append(f"{seg.sid}\t{x:.3f}\t{y:.3f}\t{z:.3f}\t{seg.d[n]:.3f}")
                data_part += "\n".join(seg_data) + "\n\n"
                number_of_points += len(seg.d)
            header = dedent(f"""\
            # MicroVisu3D file (created by {config.app_name} {config.version})
            # Number of lines   {len(net.segment)}
            # Number of points  {number_of_points}
            # Number of inter.  {len(net.cross_point_dict)}
            #
            # No\tx\ty\tz\td
            #\n""")
            out_path.write_text(header + data_part)
            return out_path

        @staticmethod
        def x3d(net, out_path='.', overwrite=False,
                color_mode='flat', color_map='jet', color=None, object_type=None, segment_list=None,
                azimuth=None, elevation=None, roll=None):
            """
            Export :class:`Qiber3D.Network` as :file:`.x3d` file.

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :param str color_mode: sets the way to color the network
                choose one of ['flat', 'fiber', 'fiber_length', 'fiber_volume', 'segment',
                'segment_length', 'segment_volume', 'fiber_segment_ratio']
            :param str color_map: name of a matplotlib colormap
            :param tuple(float) color: color if color_mode is `'flat'`
            :param str object_type: when set to `'line'` render center line
            :param tuple segment_list: limit the visualisation to certain segment (use sid)
            :param float azimuth: change camera azimuth
            :param float elevation: change camera elevation
            :param float roll: roll camera
            :return: path to saved file
            :rtype: Path
            """
            out_path = net.render.export_x3d(out_path=out_path, overwrite=overwrite,
                                             color_mode=color_mode, color_map=color_map, color=color,
                                             object_type=object_type, segment_list=segment_list,
                                             azimuth=azimuth, elevation=elevation, roll=roll)
            return out_path

        @staticmethod
        def tif(net, out_path='.',  overwrite=False, voxel_resolution=None, segment_list=None):
            """
            Export :class:`Qiber3D.Network` as :file:`.tif` image stack.

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network, if `None` show the plot.
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :param float voxel_resolution: number of voxels per unit length
            :param tuple segment_list: limit the visualisation to certain segment (use sid)
            :return: path to saved file
            :rtype: Path
            """

            out_path = net.render.export_image_stack(out_path=out_path, overwrite=overwrite,
                                                     voxel_resolution=voxel_resolution, segment_list=segment_list)

            return out_path

        @staticmethod
        def json(net, out_path='.', overwrite=False):
            """
            Export :class:`Qiber3D.Network` as :file:`.json` file.

            :param Qiber3D.Network net: network to export
            :param out_path: file or folder path where to save the network
            :type out_path: str, Path
            :param bool overwrite: allow file overwrite
            :return: path to saved file
            :rtype: Path
            """
            out_path, needs_unlink = helper.out_path_check(out_path, network=net, prefix='', suffix='.json',
                                                           overwrite=overwrite, logger=net.logger)
            if out_path is None:
                return

            def clean_value(value):
                if type(value) == np.ndarray:
                    value = value.tolist()
                elif type(value) in [np.float32, np.float64]:
                    value = float(value)
                elif isinstance(value, int):
                    value = int(value)
                return value

            data = dict(
                meta=dict(
                    created=datetime.now().isoformat(),
                    app_name=config.app_name,
                    app_version=config.version
                ),
                network=dict(
                    fiber=dict(),
                    extractor_data=None
                )
            )

            if net.extractor_data:
                data['network']['extractor_data'] = dict()
                for key, value in net.extractor_data.items():
                    if 'processing_data' == key:
                        data['network']['extractor_data'][key] = dict()
                        for pro_key in net.extractor_data[key]:
                            data['network']['extractor_data'][key][pro_key] = dict()
                            for sub_key, sub_value in net.extractor_data[key][pro_key].items():
                                data['network']['extractor_data'][key][pro_key][sub_key] = clean_value(sub_value)
                    else:
                        data['network']['extractor_data'][key] = clean_value(value)

            if net.input_file:
                data['meta']['source'] = str(net.input_file.absolute())
            else:
                data['meta']['source'] = None

            if net.name:
                data['meta']['name'] = net.name
            else:
                data['meta']['name'] = None

            for key in ['bbox', 'bbox_size', 'bbox_volume', 'center', 'volume', 'average_radius',
                        'max_radius', 'cylinder_radius', 'length']:
                data['network'][key] = clean_value(getattr(net, key))

            for fid, fiber in net.fiber.items():
                data['network']['fiber'][fid] = dict()
                for key in ['fid', 'volume', 'average_radius', 'cylinder_radius', 'length', 'sid_list']:
                    data['network']['fiber'][fid][key] = clean_value(getattr(fiber, key))

            data['network']['segment'] = dict()
            for sid, segment in net.segment.items():
                data['network']['segment'][sid] = dict()
                for key in ['sid', 'volume', 'average_radius', 'cylinder_radius', 'length',
                            'direction', 'point', 'radius']:
                    data['network']['segment'][sid][key] = clean_value(getattr(segment, key))

            out_path.write_text(json.dumps(data))
            return out_path
